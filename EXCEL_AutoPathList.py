# -*- encoding: iso-8859-15 -*-

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment  
from pathlib import Path
import tkinter
import tkinter.filedialog
import os


KiB = 1024
MiB = KiB * KiB
GiB = MiB * KiB

root = tkinter.Tk()
root.withdraw()         #use to hide tkinter window
tempdir = tkinter.filedialog.askdirectory(parent=root, initialdir=os.getcwd(), title='Please select a directory')
root_directory = Path(tempdir)

Dateiname = "Liste.xlsx"
wb = Workbook()
ws = wb.active

ws['A1'] = "Dateipfad"
ws['B1'] = "Name"
ws['C1'] = "Größe in GiB"
ws['D1'] = "Größe in GB"


ListenPfad = tempdir + "/" + Dateiname

if len(tempdir) > 0:
    print ("You chose %s" % tempdir)

i = 2
for subFolder in os.listdir(tempdir):
    t = subFolder
    subFolder = tempdir+ "/" + subFolder
    temp = sum(f.stat().st_size for f in Path(subFolder).glob('**/*') if f.is_file())    
    ws['A'+str(i)] = subFolder
    ws['B'+str(i)] = str(t)
    ws['C'+str(i)] = round(temp/GiB,3)
    ws['D'+str(i)] = round(temp/10**9,3)
    i += 1


SummenZeile = "A"+str(i+1)+":B"+str(i+1)
SumZelleGib = "C"+str(i+1)
SumZelleGb = "D"+str(i+1)

ws['A'+str(i+1)] = "Summe: "
ws[SumZelleGib] = "=Sum(C2:C"+str(i-1)+")"
ws[SumZelleGb] = "=Sum(D2:D"+str(i-1)+")"
ws.merge_cells(SummenZeile)
    
wb.save(ListenPfad)
StartEXCELFile = "start EXCEL.EXE " + ListenPfad
os.system(StartEXCELFile)
print("Liste wurde generiert! Dateipfad lautet: \n\t" + ListenPfad)




